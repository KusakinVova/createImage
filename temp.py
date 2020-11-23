filename = "123.xlsx"
row_start = 2
row_end = 200000
colum_start = 1
colum_end = 4
sheet_name = 'Price 2020-2'
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./' + filename)
# Get sheet names
# print(wb.get_sheet_names())
# Get a sheet by name
sheet = wb.get_sheet_by_name(sheet_name)

long_name_0 = 0
long_name_1 = ''
long_name_2 = ''
count_len_0 = 0
count_len_1 = 0
count_len_2 = 0

for i1 in range(row_start, row_end):
    name0 = str(sheet.cell(row=i1, column=1).value)
    name1 = str(sheet.cell(row=i1, column=2).value)
    name2 = str(sheet.cell(row=i1, column=3).value)
    if( len(name0) > count_len_0 ):
        long_name_0 = name0
        count_len_0 = len(name0)
    if( len(name1) > count_len_1 ):
        long_name_1 = name1
        count_len_1 = len(name1)
    if( len(name2) > count_len_2 ):
        long_name_2 = name2
        count_len_2 = len(name2)

print(long_name_0)
print(count_len_0)
print(long_name_1)
print(count_len_1)
print(long_name_2)
print(count_len_2)