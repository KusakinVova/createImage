filename = "./dop/1848575397.xlsx"
# filename = "123.xlsx"
row_titles = 1
use_title = False
row_start = 2
row_end = 100
colum_start = 1
colum_end = 4
sheet_name = 'Price 2020-2'
imgprefix = 'img_'
img_dir = './imgs/'
pretextqr = 'https://barkaris.ru/catalog/?q='
email = 'email: sale@barkaris.ru'


import os

import qrcode

import barcode
from barcode.writer import ImageWriter

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageFilter

def createimg(text, filename):
    # use function for create images
    # createimg("Hello\nWorld", "img")
    fnt = ImageFont.truetype('/Library/Fonts/Arial Unicode.ttf', 35)

    #new img
    im_rgba = Image.new('RGB', (700, 500), color=(255, 255, 255))

    # use image
    #img = Image.open('./main.jpeg')
    #im_rgba = img.copy()
    #im_rgba.putalpha(128)

    d = ImageDraw.Draw(im_rgba)
    d.text((50, 50), text, font=fnt, fill=(0, 0, 0))
    im_rgba.save( img_dir + filename + '.png', optimize=True, quality=25)


def createimg2(numb, name1, name2):
    # use function for create images

    if(len(name1) > 20 ):
        font_size1 = 20
        name1 = editstring(name1)
        up_name1 = 135
    else:
        font_size1 = 25
        up_name1 = 150

    if(len(name2) > 20 ):
        font_size2 = 20
        name2 = editstring(name2)
        up_name2 = 185
    else:
        font_size2 = 25
        up_name2 = 180

    fnt1 = ImageFont.truetype('/Library/Fonts/Arial Unicode.ttf', 45)
    fnt2 = ImageFont.truetype('/Library/Fonts/Arial Unicode.ttf', font_size1)
    fnt3 = ImageFont.truetype('/Library/Fonts/Arial Unicode.ttf', font_size2)
    fnt4 = ImageFont.truetype('/Library/Fonts/Arial Unicode.ttf', 25)

    #create barcode
    ean = barcode.get('code128', numb, writer=ImageWriter())
    ean.save( img_dir + 'b_' + numb, {"module_width": 0.35, "module_height": 5, "font_size": 0, "text_distance": 0}  )
    # code128.save('filename', {"module_width": 0.35, "module_height": 10, "font_size": 18, "text_distance": 1, "quiet_zone": 3})

    #new img
    im_rgba = Image.new('RGB', (800, 333), color=(255, 255, 255))
    im_rgba2 = Image.new('RGB', (790, 323), color=(0, 0, 0))
    im_rgba3 = Image.new('RGB', (457, 310), color=(255, 255, 255))
    im_rgba4 = Image.new('RGB', (457, 6), color=(0, 0, 0))
    im_rgba.paste(im_rgba2, (5, 5))
    im_rgba.paste(im_rgba3, (330, 11))
    im_rgba.paste(im_rgba4, (330, 265))

    #add bar code
    img_bar = Image.open(img_dir + 'b_' + numb+'.png')
    im_rgba.paste(img_bar, (350, 11))
    os.remove(img_dir + 'b_' + numb+'.png')

    #create qr
    qr = qrcode.QRCode(box_size=10,border=1)
    qr.add_data(pretextqr+numb)
    qr.make()
    img_qr = qr.make_image()
    im_rgba.paste(img_qr, (11,11))

    d = ImageDraw.Draw(im_rgba)
    d.text((440, 80), numb, font=fnt1, fill=(0, 0, 0))
    d.text((380, up_name1), name1, font=fnt2, fill=(0, 0, 0))
    d.text((380, up_name2), name2, font=fnt3, fill=(0, 0, 0))
    d.text((380, 275), email, font=fnt4, fill=(0, 0, 0))
    im_rgba.save( img_dir + numb + '.png', optimize=True, quality=25)



#======================================
def editstring(oldstring):
    startpoint = 0
    endpoint = 38
    maxlen = 38
    stop = 0
    newstring = ''
    while(stop == 0):
        index = oldstring.rfind(' ', startpoint, endpoint)
        newstring = newstring + oldstring[startpoint+1:index]+'\n'
        endpoint = index + maxlen
        startpoint = index+1
        if( (len(oldstring) - startpoint) < maxlen):
            stop = 1
            newstring = newstring + oldstring[startpoint:len(oldstring)]
    return newstring
#======================================

def main():

    #create dir
    if(os.path.isdir(img_dir) == False):
        os.mkdir(img_dir)

    # Load in the workbook
    wb = load_workbook('./'+filename)
    # Get sheet names
    #print(wb.get_sheet_names())
    # Get a sheet by name
    sheet = wb[sheet_name]


    for i1 in range(row_start, row_end):
        numb = str(sheet.cell(row=i1, column=1).value)
        name1 = str(sheet.cell(row=i1, column=2).value)
        name2 = str(sheet.cell(row=i1, column=3).value)

        createimg2(numb, name1, name2)
        #print(numb)

#======================================
main() # start programm
